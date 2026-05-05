import os
from openpyxl import Workbook, load_workbook
from config import InfoMaquinas
from InquirerPy import prompt
from openpyxl.styles import Font, Alignment

nDE = input("Digite o nome da empresa: ")
arquivo = f"{nDE}.xlsx"

modificar = [
    {
        "type": "list",
        "name": "modificacao",
        "message": "Deseja modificar a planilha ?",
        "choices": ["Sim", "Nao"]
    }
]

hd = [
    {
        "type": "list",
        "name": "modificacao",
        "message": "SSD ou HD ?",
        "choices": ["SSD", "HD"]
    }
]

if os.path.exists(arquivo):

    resposta_mod = prompt(modificar)

    if resposta_mod["modificacao"] == "Sim":

        resposta_hd = prompt(hd)  # movido para cá, só pergunta se for modificar

        wb = load_workbook(arquivo)
        ws = wb.active

        linha = ws.max_row + 1

        ws[f"A{linha}"] = input("Digite a etiqueta da maquina:")
        ws[f"B{linha}"] = InfoMaquinas.system.Name
        ws[f"C{linha}"] = input("Digite o nome do usuario: ")
        ws[f"D{linha}"] = input("Digite o departamento: ")
        ws[f"E{linha}"] = InfoMaquinas.tipo_pc
        ws[f"F{linha}"] = InfoMaquinas.system.Manufacturer
        ws[f"G{linha}"] = InfoMaquinas.system.Model
        ws[f"H{linha}"] = str(InfoMaquinas.nome_p)
        ws[f"I{linha}"] = str(InfoMaquinas.memoria_total) + " "+ (InfoMaquinas.tipo_ram)
        ws[f"J{linha}"] = str(InfoMaquinas.disco_total) + " " + resposta_hd["modificacao"]
        ws[f"K{linha}"] = InfoMaquinas.nome_so
        ws[f"L{linha}"] = input("Alguma Obs: ")

        wb.save(arquivo)
        print("Planilha modificada com sucesso")

    else:
        print("Operação cancelada.")

else:
    wb = Workbook()
    ws = wb.active

    resposta_hd = prompt(hd)

    ws["A1"] = nDE
    cell = ws["A1"]
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center")

    ws["A2"] = "Etiqueta"
    ws["B2"] = "Nome da maquina"
    ws["C2"] = "Usuario"
    ws["D2"] = "Departamento"
    ws["E2"] = "Gabinete"
    ws["F2"] = "Marca"
    ws["G2"] = "Modelo"
    ws["H2"] = "Processador"
    ws["I2"] = "Memoria"
    ws["J2"] = "HD"
    ws["K2"] = "Software"
    ws["L2"] = "OBS"

    linha = 3

    ws[f"A{linha}"] = input("Digite a etiqueta da maquina: ")
    ws[f"B{linha}"] = InfoMaquinas.system.Name
    ws[f"C{linha}"] = input("Digite o nome do usuario: ")
    ws[f"D{linha}"] = input("Digite o departamento: ")
    ws[f"E{linha}"] = InfoMaquinas.tipo_pc
    ws[f"F{linha}"] = InfoMaquinas.system.Manufacturer
    ws[f"G{linha}"] = InfoMaquinas.system.Model
    ws[f"H{linha}"] = str(InfoMaquinas.nome_p)
    ws[f"I{linha}"] = str(InfoMaquinas.memoria_total) + " " + (InfoMaquinas.tipo_ram)
    ws[f"J{linha}"] = str(InfoMaquinas.disco_total) + " " + resposta_hd["modificacao"]
    ws[f"K{linha}"] = InfoMaquinas.nome_so
    ws[f"L{linha}"] = input("Alguma Obs: ")

    wb.save(arquivo)
    print("Planilha criada com sucesso")