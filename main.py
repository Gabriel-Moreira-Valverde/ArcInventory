import os
from openpyxl import Workbook, load_workbook
from config import InfoMaquinas
from InquirerPy import prompt
from openpyxl.styles import Font, Alignment

# =========================
# 📌 Entrada do usuário
# =========================
nome_empresa = input("Digite o nome da empresa: ")
arquivo = f"{nome_empresa}.xlsx"

# =========================
# 📋 Perguntas (menus)
# =========================
modificar = [
    {
        "type": "list",
        "name": "modificacao",
        "message": "Deseja modificar a planilha ?",
        "choices": ["Sim", "Nao"]
    }
]

hd = [
    {
        "type": "list",
        "name": "tipo_hd",
        "message": "SSD, HD ou NVME?",
        "choices": ["SSD", "HD", "NVME"]
    }
]

# =========================
# 🧠 Função principal
# Preenche uma linha da planilha
# =========================
def preencher_linha(ws, linha, tipo_hd):
    ws[f"A{linha}"] = input("Digite a etiqueta da maquina: ")
    ws[f"B{linha}"] = InfoMaquinas.system.Name
    ws[f"C{linha}"] = input("Digite o nome do usuario: ")
    ws[f"D{linha}"] = input("Digite o departamento: ")
    ws[f"E{linha}"] = InfoMaquinas.tipo_pc
    ws[f"F{linha}"] = InfoMaquinas.system.Manufacturer
    ws[f"G{linha}"] = InfoMaquinas.system.Model
    ws[f"H{linha}"] = str(InfoMaquinas.nome_processador)
    ws[f"I{linha}"] = f"{InfoMaquinas.memoria_total} {InfoMaquinas.tipo_ram}"
    ws[f"J{linha}"] = f"{InfoMaquinas.disco_total} {tipo_hd}"
    ws[f"K{linha}"] = InfoMaquinas.nome_software
    ws[f"L{linha}"] = input("Alguma Obs: ")

# =========================
# 📂 Caso o arquivo já exista
# =========================
if os.path.exists(arquivo):

    # Pergunta se quer editar
    resposta_mod = prompt(modificar)

    # Pergunta tipo de HD
    resposta_hd = prompt(hd)

    if resposta_mod["modificacao"] == "Sim":

        # Abre planilha existente
        wb = load_workbook(arquivo)
        ws = wb.active

        # Próxima linha disponível
        linha = ws.max_row + 1

        # Preenche dados na planilha
        preencher_linha(ws, linha, resposta_hd["tipo_hd"])

        wb.save(arquivo)
        print("Planilha modificada com sucesso")

    else:
        print("Operação cancelada.")

# =========================
# 📄 Caso o arquivo NÃO exista
# =========================
else:
    wb = Workbook()
    ws = wb.active

    resposta_hd = prompt(hd)

    # =========================
    # 🏷️ Título da planilha
    # =========================
    ws["A1"] = nome_empresa
    cell = ws["A1"]
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center")

    # =========================
    # 📊 Cabeçalho da tabela
    # =========================
    ws["A2"] = "Etiqueta"
    ws["B2"] = "Nome da maquina"
    ws["C2"] = "Usuario"
    ws["D2"] = "Departamento"
    ws["E2"] = "Gabinete"
    ws["F2"] = "Marca"
    ws["G2"] = "Modelo"
    ws["H2"] = "Processador"
    ws["I2"] = "Memoria"
    ws["J2"] = "HD"
    ws["K2"] = "Software"
    ws["L2"] = "OBS"

    # =========================
    # ➕ Primeira linha de dados
    # =========================
    linha = 3
    preencher_linha(ws, linha, resposta_hd["tipo_hd"])

    wb.save(arquivo)
    print("Planilha criada com sucesso")